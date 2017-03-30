import openpyxl
import pickle
import itertools

def add(one, two):

    return [one[a] + two[a] for a in range(14)]

def grpSort(matrix, key):

    return itertools.groupby(sorted(matrix, key= lambda s: s[key]), key= lambda s: s[key])

wb = openpyxl.load_workbook('D:\MyDrive\FreeThrows.xlsx')
postseason = pickle.load(open( "post.p.p", "rb" ))
regularseason = pickle.load(open( "reg.p", "rb" ))
list = []
list.extend(postseason)
list.extend(regularseason)
length = len(list)

templist = list
#def combine():
sortedL = []



print "sorted\n\n"

for a in grpSort(templist, "season"):
    for b in grpSort(a[1], "seasonType"):
        for d in grpSort(b[1], "opp"):
            for e in grpSort(d[1], "location"):
                totals = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
                for game in e[1]:
                    totals = add(totals, game["stats"])
                sortedL.append({"season" : game["season"], "seasonType" : game["seasonType"],
                                "location" : game["location"], "opp" : game["opp"], "stats" : totals})


sheet = wb.get_sheet_by_name('LeBron James')
sheet.cell(row = 1, column = 1).value = "Season"
sheet.cell(row = 1, column = 2).value = "S. Type"
sheet.cell(row = 1, column = 3).value = "Opponent"
sheet.cell(row = 1, column = 4).value = "Location"
sheet.cell(row = 1, column = 5).value = "Make 1"
sheet.cell(row = 1, column = 6).value = "Miss 1"
sheet.cell(row = 1, column = 7).value = "Make 2"
sheet.cell(row = 1, column = 8).value = "Miss 2"
sheet.cell(row = 1, column = 9).value = "Make 3"
sheet.cell(row = 1, column = 10).value = "Miss 3"
sheet.cell(row = 1, column = 11).value = "Make Tech."
sheet.cell(row = 1, column = 12).value = "Miss Tech"
sheet.cell(row = 1, column = 13).value = "Make F.1"
sheet.cell(row = 1, column = 14).value = "Miss F.1"
sheet.cell(row = 1, column = 15).value = "Make F.2"
sheet.cell(row = 1, column = 16).value = "Miss F.2"
sheet.cell(row = 1, column = 17).value = "Make CP"
sheet.cell(row = 1, column = 18).value = "Miss CP"

for i in range(len(sortedL)):
    sheet.cell(row = i+2, column = 1).value = sortedL[i]["season"]
    sheet.cell(row = i+2, column = 2).value = sortedL[i]["seasonType"]
    sheet.cell(row = i+2, column = 3).value = sortedL[i]["opp"]
    sheet.cell(row = i+2, column = 4).value = sortedL[i]["location"]
    for j in range(14):
        sheet.cell(row = i+2, column = j + 5).value = sortedL[i]["stats"][j]





wb.save('D:\MyDrive\FreeThrowsEdited.xlsx')
